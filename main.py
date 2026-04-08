"""
Rose Glass Sales Platform — Backend
====================================
Scout agents, CERATA ranking engine, chat API.

Supabase project: xrzycmvpqohxxlhnorpt (rose-glass-sales)
Deployment: Railway

Author: Christopher MacGregor bin Joseph
ROSE Corp / MacGregor Holding Company
"""

import os
import json
import csv
import io
import math
import asyncio
import logging
import secrets
from datetime import datetime, timezone
from typing import Dict, List, Optional, Any

from fastapi import FastAPI, HTTPException, Request, UploadFile, File, Header
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import httpx

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("rose-glass-sales")

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://xrzycmvpqohxxlhnorpt.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
PORT = int(os.environ.get("PORT", 8000))

HEADERS_SB = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

app = FastAPI(title="Rose Glass Sales Platform", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ═══════════════════════════════════════════════════════════
# CERATA V2 BRIDGE — Computational dimensional analysis
# ═══════════════════════════════════════════════════════════

class CERATABridge:
    """
    Computational Rose Glass analysis for buying signals.
    No LLM calls — pure math on signal text.

    Dimensions (CRM calibration):
    - Ψ: Intent coherence — need/solution match
    - ρ: Decision authority — can they buy?
    - q: Urgency/activation — bio-optimized via Michaelis-Menten
    - f: Fit/belonging — ICP alignment

    C = Ψ + (ρ × Ψ) + q_opt + (f × Ψ) + coupling
    q_opt = q / (Km + q + q²/Ki)
    """

    KM = 0.20
    KI = 0.80
    COUPLING_STRENGTH = 0.15

    PSI_SIGNALS = {
        "looking for": 0.7, "need a solution": 0.8, "evaluating": 0.7,
        "searching for": 0.6, "want to implement": 0.8, "interested in": 0.5,
        "exploring options": 0.6, "considering": 0.5, "planning to": 0.7,
        "ready to": 0.8, "seeking": 0.6, "require": 0.7, "must have": 0.8,
        "pain point": 0.7, "challenge": 0.5, "struggling with": 0.7,
        "frustrated with": 0.6, "current solution failing": 0.9,
        "looking to switch": 0.8, "replacing": 0.8, "upgrading": 0.7,
        "scaling": 0.6, "growing": 0.5, "expanding": 0.5,
        "client management": 0.7, "patient engagement": 0.7,
        "admissions": 0.6, "intake": 0.6, "referral": 0.6,
        "census": 0.7, "bed count": 0.6, "occupancy": 0.7,
        "marketing roi": 0.7, "lead tracking": 0.8,
    }

    RHO_SIGNALS = {
        "owner": 0.9, "ceo": 0.9, "founder": 0.9, "president": 0.8,
        "chief": 0.8, "vp": 0.7, "vice president": 0.7, "director": 0.6,
        "head of": 0.7, "c-suite": 0.9, "decision maker": 0.8,
        "budget authority": 0.9, "budget holder": 0.9,
        "approved budget": 0.9, "allocated funds": 0.8,
        "signing authority": 0.9, "can approve": 0.8,
        "years of experience": 0.5, "previously implemented": 0.6,
        "managed": 0.5, "led": 0.5, "built": 0.6,
        "multiple locations": 0.6, "multi-site": 0.6,
    }

    Q_SIGNALS = {
        "urgent": 0.8, "immediately": 0.9, "asap": 0.9,
        "this quarter": 0.7, "this month": 0.8, "this week": 0.9,
        "deadline": 0.7, "time-sensitive": 0.8, "critical": 0.7,
        "losing": 0.8, "bleeding": 0.9, "hemorrhaging": 0.9,
        "crisis": 0.8, "emergency": 0.9, "can't wait": 0.8,
        "falling behind": 0.7, "competitors": 0.6,
        "regulatory": 0.7, "compliance deadline": 0.8,
        "audit": 0.6, "accreditation": 0.7, "jcaho": 0.7, "carf": 0.7,
        "hiring": 0.5, "expanding": 0.5, "new location": 0.6,
        "just opened": 0.6, "launching": 0.6, "recently funded": 0.7,
    }

    F_SIGNALS = {
        # Buying ecosystem signals — industry agnostic
        "crm": 0.7, "hubspot": 0.8, "salesforce": 0.6, "pipedrive": 0.6,
        "sales team": 0.7, "sales ops": 0.7, "revenue operations": 0.7,
        "lead generation": 0.8, "pipeline": 0.7, "quota": 0.6,
        "b2b": 0.6, "enterprise sales": 0.7, "smb": 0.5,
        "saas": 0.7, "software": 0.5, "platform": 0.5,
        "growth": 0.5, "scaling": 0.6, "expansion": 0.6,
        "series a": 0.6, "series b": 0.7, "funded": 0.6, "raised": 0.6,
        "ipo": 0.5, "revenue": 0.5, "arr": 0.7, "mrr": 0.7,
        "startup": 0.4, "growing": 0.5, "hiring": 0.5,
        "multi-site": 0.6, "multiple locations": 0.6,
        "partnership": 0.5, "integration": 0.5, "api": 0.4,
        # Company maturity signals
        "inc. 5000": 0.7, "inc 5000": 0.7, "fast-growing": 0.6,
        "market leader": 0.6, "industry leader": 0.6,
        # Decision infrastructure
        "vp of sales": 0.7, "sales leader": 0.7, "head of sales": 0.7,
        "business development": 0.6, "account executive": 0.5,
    }

    @classmethod
    def analyze(cls, buying_signals="", web_signals="", linkedin_summary="",
                title="", company_industry="", company_size=0,
                company_description="", icp=None, lead_data=None):
        combined = " ".join(filter(None, [
            buying_signals, web_signals, linkedin_summary,
            title, company_industry, company_description,
        ])).lower()

        if not combined.strip():
            return dict(psi_intent=0, rho_authority=0, q_urgency=0, q_optimized=0,
                        f_fit=0, coherence_score=0, qualification_tier="unscored",
                        dimensional_fractures="No signal data", rank_score=0)

        psi = cls._extract(combined, cls.PSI_SIGNALS)
        rho = cls._extract(combined, cls.RHO_SIGNALS)
        rho = cls._boost_title(rho, title)
        q_raw = cls._extract(combined, cls.Q_SIGNALS)
        f = cls._extract(combined, cls.F_SIGNALS)
        f = cls._boost_industry(f, company_industry)

        # Apply ICP boost if user has defined a profile
        if icp and lead_data:
            psi, rho, f = cls._boost_icp(psi, rho, f, lead_data, icp)

        q_opt = cls._mm(q_raw)
        coupling = cls.COUPLING_STRENGTH * rho * psi
        coherence = psi + (rho * psi) + q_opt + (f * psi) + coupling

        tier = cls._tier(coherence, rho, q_opt, f)
        fractures = cls._fractures(psi, rho, q_raw, f)
        rank = coherence * 0.4 + rho * 0.25 + psi * 0.20 + q_opt * 0.10 + f * 0.05

        return dict(
            psi_intent=round(psi, 4), rho_authority=round(rho, 4),
            q_urgency=round(q_raw, 4), q_optimized=round(q_opt, 4),
            f_fit=round(f, 4), coherence_score=round(coherence, 4),
            qualification_tier=tier, dimensional_fractures=fractures,
            rank_score=round(rank, 4),
        )

    @classmethod
    def _extract(cls, text, signals):
        matches = sorted([w for s, w in signals.items() if s in text], reverse=True)
        return min(1.0, sum(matches[:3]) / len(matches[:3])) if matches else 0.0

    @classmethod
    def _boost_title(cls, rho, title):
        if not title: return rho
        t = title.lower()
        if any(x in t for x in ["owner","ceo","founder","president","co-founder"]): return min(1.0, rho+0.4)
        if any(x in t for x in ["chief","cmo","cto","cfo","coo","cro"]): return min(1.0, rho+0.35)
        if any(x in t for x in ["vp","vice president","svp","evp"]): return min(1.0, rho+0.25)
        if any(x in t for x in ["director","head of"]): return min(1.0, rho+0.15)
        if any(x in t for x in ["manager","lead"]): return min(1.0, rho+0.05)
        return rho

    @classmethod
    def _boost_industry(cls, f, industry):
        """Boost fit based on company having a sales-oriented industry, not a specific vertical."""
        if not industry: return f
        ind = industry.lower()
        # Companies in sales-heavy industries get a fit boost — they understand buying tools
        if any(x in ind for x in ["software","technology","saas","information tech"]): return min(1.0, f+0.15)
        if any(x in ind for x in ["sales","marketing","consulting","professional service"]): return min(1.0, f+0.20)
        if any(x in ind for x in ["health","medical","pharma","biotech"]): return min(1.0, f+0.10)
        if any(x in ind for x in ["financial","insurance","real estate"]): return min(1.0, f+0.10)
        return f

    @classmethod
    def _boost_icp(cls, psi, rho, f, lead_data, icp):
        """Boost dimensions based on user-defined ICP match."""
        if not icp:
            return psi, rho, f
        
        title = (lead_data.get("title") or "").lower()
        industry = (lead_data.get("company_industry") or "").lower()
        region = (lead_data.get("region") or "").lower()
        company_size = lead_data.get("company_size") or 0
        text = " ".join(filter(None, [title, industry, lead_data.get("company_description",""), lead_data.get("buying_signals","")])).lower()
        
        boosts = 0
        
        # Title match
        for t in icp.get("target_titles", []):
            if t.lower() in title:
                rho = min(1.0, rho + 0.15)
                boosts += 1
                break
        
        # Industry match
        for ind in icp.get("target_industries", []):
            if ind.lower() in industry:
                f = min(1.0, f + 0.20)
                boosts += 1
                break
        
        # Company size match
        size_min = icp.get("target_company_size_min")
        size_max = icp.get("target_company_size_max")
        if company_size and size_min and size_max:
            if size_min <= company_size <= size_max:
                f = min(1.0, f + 0.10)
                boosts += 1
        
        # Region match
        for r in icp.get("target_regions", []):
            if r.lower() in region:
                f = min(1.0, f + 0.05)
                boosts += 1
                break
        
        # Keyword match
        for kw in icp.get("target_keywords", []):
            if kw.lower() in text:
                psi = min(1.0, psi + 0.05)
                boosts += 1
        
        # Exclusion match — penalize
        for kw in icp.get("exclude_keywords", []):
            if kw.lower() in text:
                psi = max(0, psi - 0.3)
                f = max(0, f - 0.3)
                break
        
        return psi, rho, f

    @classmethod
    def _mm(cls, q):
        return q / (cls.KM + q + q*q/cls.KI) if q > 0 else 0.0

    @classmethod
    def _tier(cls, c, rho, q_opt, f):
        if c < 0.5 or f < 0.2: return "disqualified"
        if c >= 2.5 and rho >= 0.6 and q_opt >= 0.15: return "hot"
        if c >= 1.5: return "warm"
        if c >= 0.5: return "cold"
        return "disqualified"

    @classmethod
    def _fractures(cls, psi, rho, q, f):
        fr = []
        if q > 0.6 and rho < 0.3: fr.append("High urgency + low authority: may lack buying power")
        if psi > 0.6 and rho < 0.3: fr.append("Clear intent + low authority: champion without budget — find their boss")
        if rho > 0.6 and psi < 0.3: fr.append("High authority + low intent: decision maker not yet activated")
        if f > 0.6 and psi < 0.3: fr.append("Strong fit + no intent: perfect ICP but no buying motion")
        if q > 0.7 and psi < 0.4: fr.append("Urgency without clarity: activation without defined need")
        return "; ".join(fr) if fr else "No fractures detected"


# ═══════════════════════════════════════════════════════════
# SCOUT AGENT
# ═══════════════════════════════════════════════════════════

SCOUT_MODEL = os.environ.get("SCOUT_MODEL", "claude-haiku-4-5-20251001")
PERCEPTION_SEED_FILE_ID = os.environ.get("PERCEPTION_SEED_FILE_ID", "file_011CZpgkQXaWG5Bgi1tyavRq")  # Haiku for bulk scouts, override with SCOUT_MODEL env var

class ScoutAgent:
    SYSTEM = """You are an elite sales intelligence scout. You research ANY company in ANY industry.

Your mission: Given a lead, search the web thoroughly for BUYING SIGNALS that indicate this person or company
is in a growth phase, has budget, has pain points, or is actively seeking solutions. You are industry-agnostic —
an aerospace VP is just as valuable as a healthcare VP if the signals are strong.

NEVER judge whether a lead "fits" a particular industry. Your ONLY job is to find buying signals.
Do NOT output warnings about industry mismatch. Every lead is valid.

SEARCH STRATEGY (MAX 3 SEARCHES — be efficient):
1. Search "[company name] news funding hiring 2025 2026" — get company signals in one query
2. Search "[person name] [company]" — get their role, activity, authority
3. Search "[company name] jobs" — only if first two searches lack growth signals
DO NOT search more than 3 times. Combine keywords to get maximum signal per search.

SIGNAL CATEGORIES (be specific with dates and sources):
- GROWTH: New locations, hiring, expansion, new products, market entry
- PAIN POINTS: Operational challenges, turnover, scaling issues, tech debt
- TECHNOLOGY: Current tech stack, recent implementations, RFPs, vendor evaluations
- FINANCIAL: Funding rounds, revenue growth, profitability signals, M&A activity
- COMPETITIVE: Market position, competitive pressures, differentiation efforts
- LEADERSHIP: New hires in sales/BD/ops roles, board changes, strategic pivots
- AUTHORITY: This person's decision-making power, team size, budget control

OUTPUT FORMAT: Plain text paragraphs organized by signal category.
Include dates, sources, and confidence level for each signal.
If nothing found, say "No significant buying signals detected" — do NOT fabricate."""

    @classmethod
    async def scout_lead(cls, lead: Dict, sales_lens: Dict = None) -> Dict[str, str]:
        name = lead.get("full_name", "")
        company = lead.get("company", "")
        title = lead.get("title", "")
        linkedin = lead.get("linkedin_profile_url") or lead.get("linkedin", "")
        domain = lead.get("company_domain") or lead.get("domain", "")
        industry = lead.get("company_industry", "")
        region = lead.get("region", "")
        description = lead.get("company_description", "")

        prompt = f"""Research buying signals for this lead:

Name: {name}
Title: {title}
Company: {company}
Industry: {industry}
Region: {region}
Domain: {domain}
LinkedIn: {linkedin}
Company description: {description[:200] if description else 'N/A'}

Search for: recent news about {company}, job postings at {company}, {name} professional activity,
and any growth/pain/funding/hiring signals. Research this lead's industry context."""

        # Inject sales lens context if available
        if sales_lens and sales_lens.get("product_name"):
            lens_inject = f"""

SALES CONTEXT — Frame your analysis through this lens:
The user sells: {sales_lens.get('product_name', '')} — {sales_lens.get('product_description', '')}
Key value: {', '.join(sales_lens.get('value_props', []))}
{('NEVER frame as: ' + ', '.join(sales_lens.get('not_this', []))) if sales_lens.get('not_this') else ''}
Look for signals that indicate this lead could benefit from or would buy this specific product.
Frame buying signals in terms of how they relate to what the user sells."""
            prompt += lens_inject

        try:
            # Build message content — attach perception seed if available
            msg_content = []
            if PERCEPTION_SEED_FILE_ID:
                msg_content.append({"type": "document", "source": {"type": "file", "file_id": PERCEPTION_SEED_FILE_ID}})
            msg_content.append({"type": "text", "text": prompt})

            async with httpx.AsyncClient(timeout=90) as client:
                resp = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "x-api-key": ANTHROPIC_API_KEY,
                        "anthropic-version": "2023-06-01",
                        "anthropic-beta": "files-api-2025-04-14",
                        "Content-Type": "application/json",
                    },
                    json={
                        "model": SCOUT_MODEL,
                        "max_tokens": int(os.environ.get("SCOUT_MAX_TOKENS", "1024")),
                        "system": cls.SYSTEM,
                        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 3}],
                        "messages": [{"role": "user", "content": msg_content}],
                    },
                )
                resp.raise_for_status()
                data = resp.json()

            texts = [b["text"] for b in data.get("content", []) if b.get("type") == "text"]
            return {
                "buying_signals": "\n".join(texts) or "No signals found",
                "web_signals": f"Scouted by {SCOUT_MODEL} at {datetime.now(timezone.utc).isoformat()}",
                "linkedin_summary": f"LinkedIn: {linkedin}" if linkedin else "",
            }
        except Exception as e:
            logger.error(f"Scout failed for {name}: {e}")
            return {"buying_signals": f"Scout error: {str(e)[:200]}", "web_signals": "", "linkedin_summary": ""}


# ═══════════════════════════════════════════════════════════
# API ROUTES
# ═══════════════════════════════════════════════════════════

@app.get("/health")
async def health():
    return {"status": "ok", "service": "rose-glass-sales"}


# ─── Auth ─────────────────────────────────────────────────

class RegisterRequest(BaseModel):
    username: str
    password: str
    display_name: Optional[str] = None
    company_name: Optional[str] = None
    industry: Optional[str] = None

class LoginRequest(BaseModel):
    username: str
    password: str

class ProfileUpdate(BaseModel):
    display_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    company_name: Optional[str] = None
    industry: Optional[str] = None

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

class UsernameChange(BaseModel):
    new_username: str
    password: str


async def get_current_user(authorization: str = Header(None)) -> Dict:
    """Extract user from Bearer token."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Not authenticated")
    token = authorization.replace("Bearer ", "")
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/sessions?token=eq.{token}&select=user_id,expires_at",
            headers=HEADERS_SB)
        sessions = resp.json()
    if not sessions:
        raise HTTPException(401, "Invalid session")
    session = sessions[0]
    if session["expires_at"] and session["expires_at"] < datetime.now(timezone.utc).isoformat():
        raise HTTPException(401, "Session expired")
    return {"user_id": session["user_id"]}


@app.post("/api/auth/register")
async def register(req: RegisterRequest):
    if len(req.username) < 3:
        raise HTTPException(400, "Username must be at least 3 characters")
    if len(req.password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")
    async with httpx.AsyncClient(timeout=15) as client:
        # Check if username taken
        check = await client.get(
            f"{SUPABASE_URL}/rest/v1/users?username=eq.{req.username}&select=id",
            headers=HEADERS_SB)
        if check.json():
            raise HTTPException(400, "Username already taken")
        # Create user (hash password in SQL via pgcrypto)
        resp = await client.post(
            f"{SUPABASE_URL}/rest/v1/rpc/create_user",
            headers=HEADERS_SB,
            json={"p_username": req.username, "p_password": req.password,
                  "p_display_name": req.display_name, "p_company_name": req.company_name,
                  "p_industry": req.industry})
        if resp.status_code >= 400:
            raise HTTPException(400, f"Registration failed: {resp.text[:200]}")
        user_id = resp.json()
    # Auto-login
    token = secrets.token_urlsafe(48)
    expires = (datetime.now(timezone.utc) + __import__("datetime").timedelta(days=30)).isoformat()
    async with httpx.AsyncClient(timeout=15) as client:
        await client.post(f"{SUPABASE_URL}/rest/v1/sessions", headers=HEADERS_SB,
            json={"user_id": user_id, "token": token, "expires_at": expires})
    return {"token": token, "user_id": user_id, "username": req.username}


@app.post("/api/auth/login")
async def login(req: LoginRequest):
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(
            f"{SUPABASE_URL}/rest/v1/rpc/verify_login",
            headers=HEADERS_SB,
            json={"p_username": req.username, "p_password": req.password})
        if resp.status_code >= 400 or not resp.json():
            raise HTTPException(401, "Invalid username or password")
        user_id = resp.json()
    token = secrets.token_urlsafe(48)
    expires = (datetime.now(timezone.utc) + __import__("datetime").timedelta(days=30)).isoformat()
    async with httpx.AsyncClient(timeout=15) as client:
        await client.post(f"{SUPABASE_URL}/rest/v1/sessions", headers=HEADERS_SB,
            json={"user_id": user_id, "token": token, "expires_at": expires})
        # Get user profile
        profile = await client.get(
            f"{SUPABASE_URL}/rest/v1/users?id=eq.{user_id}&select=id,username,display_name,email,phone,company_name,industry,avatar_url",
            headers=HEADERS_SB)
    return {"token": token, "user": profile.json()[0] if profile.json() else {"id": user_id}}


@app.get("/api/auth/me")
async def get_me(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/users?id=eq.{user['user_id']}&select=id,username,display_name,email,phone,company_name,industry,avatar_url,settings",
            headers=HEADERS_SB)
    if not resp.json():
        raise HTTPException(404, "User not found")
    return resp.json()[0]


@app.patch("/api/auth/profile")
async def update_profile(update: ProfileUpdate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    data = {k: v for k, v in update.dict().items() if v is not None}
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.patch(
            f"{SUPABASE_URL}/rest/v1/users?id=eq.{user['user_id']}",
            headers=HEADERS_SB, json=data)
    return {"updated": True}


@app.post("/api/auth/change-password")
async def change_password(req: PasswordChange, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    async with httpx.AsyncClient(timeout=15) as client:
        verify = await client.post(f"{SUPABASE_URL}/rest/v1/rpc/verify_password",
            headers=HEADERS_SB,
            json={"p_user_id": user["user_id"], "p_password": req.current_password})
        if not verify.json():
            raise HTTPException(400, "Current password is incorrect")
        await client.post(f"{SUPABASE_URL}/rest/v1/rpc/update_password",
            headers=HEADERS_SB,
            json={"p_user_id": user["user_id"], "p_new_password": req.new_password})
    return {"updated": True}


@app.post("/api/auth/change-username")
async def change_username(req: UsernameChange, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    async with httpx.AsyncClient(timeout=15) as client:
        verify = await client.post(f"{SUPABASE_URL}/rest/v1/rpc/verify_password",
            headers=HEADERS_SB,
            json={"p_user_id": user["user_id"], "p_password": req.password})
        if not verify.json():
            raise HTTPException(400, "Password is incorrect")
        resp = await client.patch(
            f"{SUPABASE_URL}/rest/v1/users?id=eq.{user['user_id']}",
            headers=HEADERS_SB, json={"username": req.new_username})
        if resp.status_code >= 400:
            raise HTTPException(400, "Username taken or invalid")
    return {"updated": True, "username": req.new_username}


@app.post("/api/auth/logout")
async def logout(authorization: str = Header(None)):
    if authorization and authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "")
        async with httpx.AsyncClient(timeout=15) as client:
            await client.delete(f"{SUPABASE_URL}/rest/v1/sessions?token=eq.{token}", headers=HEADERS_SB)
    return {"logged_out": True}


# ─── Background Scout Worker ──────────────────────────────

@app.post("/api/worker/scout")
async def worker_scout(batch_size: int = 10):
    """Background worker: scouts unscouted leads, then auto-ranks.
    Called by Railway cron at 6am and 2pm, or manually.
    Processes batch_size leads per invocation."""

    async with httpx.AsyncClient(timeout=30) as client:
        # Get unscouted leads, prioritized by rank_score (warm leads first)
        resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/leads?buying_signals=is.null&order=rank_score.desc.nullslast&limit={batch_size}",
            headers=HEADERS_SB)
        resp.raise_for_status()
        leads = resp.json()

    if not leads:
        return {"message": "All leads scouted", "remaining": 0}

    # Count remaining
    async with httpx.AsyncClient(timeout=30) as client:
        count_resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/leads?buying_signals=is.null&select=id",
            headers={**HEADERS_SB, "Prefer": "count=exact", "Range-Unit": "items", "Range": "0-0"})
        remaining = int(count_resp.headers.get("content-range", "0/0").split("/")[-1])

    # Scout one at a time (sequential to avoid rate limits)
    scouted = 0
    ranked = 0
    for lead in leads:
        try:
            signals = await ScoutAgent.scout_lead(lead)
            now = datetime.now(timezone.utc).isoformat()

            async with httpx.AsyncClient(timeout=30) as client:
                # Write scout signals
                await client.patch(
                    f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}",
                    headers=HEADERS_SB,
                    json={**signals, "buying_signals_updated_at": now,
                          "web_signals_updated_at": now, "updated_at": now})

                # Auto-rank with fresh data
                lead_resp = await client.get(
                    f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}&select=id,buying_signals,web_signals,linkedin_summary,title,company_industry,company_size,company_description",
                    headers=HEADERS_SB)
                if lead_resp.status_code == 200 and lead_resp.json():
                    ld = lead_resp.json()[0]
                    analysis = CERATABridge.analyze(
                        buying_signals=ld.get("buying_signals",""),
                        web_signals=ld.get("web_signals",""),
                        linkedin_summary=ld.get("linkedin_summary",""),
                        title=ld.get("title",""),
                        company_industry=ld.get("company_industry",""),
                        company_size=ld.get("company_size") or 0,
                        company_description=ld.get("company_description",""))
                    await client.patch(
                        f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}",
                        headers=HEADERS_SB,
                        json={**analysis, "ranked_at": now, "updated_at": now})
                    ranked += 1

            scouted += 1
            logger.info(f"Scouted [{scouted}/{len(leads)}]: {lead.get('full_name','')} at {lead.get('company','')}")

        except Exception as e:
            logger.error(f"Worker scout failed for {lead.get('full_name','')}: {e}")
            continue

    return {
        "scouted": scouted,
        "ranked": ranked,
        "remaining": remaining - scouted,
        "model": SCOUT_MODEL,
    }

# ─── CSV Upload ────────────────────────────────────────

WIZA_COLUMNS = [
    "email","full_name","first_name","last_name","title","locality","region",
    "location","linkedin","domain",
    "phone_number1","phone_number2","phone_number3","mobile_phone1","other_phone1",
    "personal_email1","company","linkedin_profile_url","company_domain",
    "company_industry","company_subindustry","company_size","company_size_range",
    "company_founded","company_revenue","company_funding","company_type",
    "company_linkedin","company_twitter","company_facebook","company_description",
    "company_last_funding_round","company_last_funding_amount","company_last_funding_at",
    "company_location","company_street","company_locality","company_region",
    "company_country","company_postal_code","other_work_emails","profile_url",
]

INT_FIELDS = {"company_size", "company_founded"}
STR_FIELDS = {"phone_number1","phone_number2","phone_number3","mobile_phone1","other_phone1","company_postal_code"}

@app.post("/api/upload-csv")
async def upload_csv(file: UploadFile = File(...), authorization: str = Header(None)):
    user = await get_current_user(authorization)
    """Upload Wiza leads from CSV, XLSX, or Numbers files."""
    fname = file.filename.lower()
    allowed = (".csv", ".xlsx", ".xls", ".numbers")
    if not any(fname.endswith(ext) for ext in allowed):
        raise HTTPException(400, f"Accepted formats: CSV, XLSX, Numbers. Got: {file.filename}")

    raw = await file.read()

    if fname.endswith(".numbers"):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".numbers", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = tmp.name
        try:
            from numbers_parser import Document
            doc = Document(tmp_path)
            table = doc.sheets[0].tables[0]
            headers_raw = [str(table.cell(0, c).value or f"col_{c}") for c in range(table.num_cols)]
            data_rows = []
            for r in range(1, table.num_rows):
                row = {}
                for c in range(table.num_cols):
                    val = table.cell(r, c).value
                    row[headers_raw[c]] = str(val) if val is not None else ""
                data_rows.append(row)
            reader_fieldnames = headers_raw
            reader_rows = data_rows
        except Exception as e:
            raise HTTPException(400, f"Could not parse .numbers file: {str(e)[:200]}. Try exporting as CSV from Apple Numbers (File > Export To > CSV).")
        finally:
            os.unlink(tmp_path)

    elif fname.endswith((".xlsx", ".xls")):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = tmp.name
        try:
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers_raw = [str(h or f"col_{i}") for i, h in enumerate(next(rows_iter))]
            data_rows = []
            for row in rows_iter:
                data_rows.append({headers_raw[i]: str(v) if v is not None else "" for i, v in enumerate(row)})
        except ImportError:
            raise HTTPException(400, "XLSX format not supported on this server. Export as CSV first.")
        finally:
            os.unlink(tmp_path)
        reader_fieldnames = headers_raw
        reader_rows = data_rows

    else:
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        reader_fieldnames = reader.fieldnames or []
        reader_rows = list(reader)

    # Validate columns
    headers = [h.strip().lower().replace(" ", "_") for h in reader_fieldnames]
    matched = [h for h in headers if h in WIZA_COLUMNS]
    if len(matched) < 5:
        raise HTTPException(400, f"File doesn't match Wiza format. Found columns: {headers[:10]}")

    rows = []
    for row in reader_rows:
        clean = {}
        for k, v in row.items():
            key = k.strip().lower().replace(" ", "_")
            if key not in WIZA_COLUMNS:
                continue
            val = (v or "").strip()
            if not val or val.lower() == "none":
                clean[key] = None
            elif key in INT_FIELDS:
                try: clean[key] = int(float(val))
                except: clean[key] = None
            elif key in STR_FIELDS:
                clean[key] = str(val).split(".")[0] if "." in str(val) else str(val)
            else:
                clean[key] = val
        if clean.get("full_name") or clean.get("first_name") or clean.get("email"):
            rows.append(clean)

    # Normalize Wiza column variants
    for row in rows:
        # Merge first_name + last_name → full_name
        if not row.get("full_name") and (row.get("first_name") or row.get("last_name")):
            row["full_name"] = f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()
        # location → locality + region
        if not row.get("locality") and row.get("location"):
            parts = row["location"].split(",", 1)
            row["locality"] = parts[0].strip() if parts else row["location"]
            row["region"] = parts[1].strip() if len(parts) > 1 else None
        # Clean out non-DB columns before insert
        for extra in ("first_name", "last_name", "location"):
            row.pop(extra, None)

    if not rows:
        raise HTTPException(400, "No valid rows found")

    # Normalize: ensure all rows have identical keys (Supabase batch requires this)
    all_keys = set()
    for row in rows:
        all_keys.update(row.keys())
    for row in rows:
        for key in all_keys:
            if key not in row:
                row[key] = None

    # Create import batch
    async with httpx.AsyncClient(timeout=30) as client:
        batch_resp = await client.post(f"{SUPABASE_URL}/rest/v1/import_batches", headers=HEADERS_SB,
            json={"filename": file.filename, "row_count": len(rows), "status": "processing", "user_id": user["user_id"]})
        batch_resp.raise_for_status()
        batch_id = batch_resp.json()[0]["id"]

        inserted = 0
        dupes = 0
        for i in range(0, len(rows), 50):
            chunk = rows[i:i+50]
            for r in chunk:
                r["import_batch_id"] = batch_id
                r["user_id"] = user["user_id"]
            resp = await client.post(
                f"{SUPABASE_URL}/rest/v1/leads",
                headers={**HEADERS_SB, "Prefer": "return=representation,resolution=merge-duplicates"},
                json=chunk)
            if resp.status_code < 300:
                inserted += len(chunk)
            else:
                logger.warning(f"Insert chunk failed: {resp.text[:200]}")

        await client.patch(f"{SUPABASE_URL}/rest/v1/import_batches?id=eq.{batch_id}", headers=HEADERS_SB,
            json={"status": "complete", "row_count": inserted})

    return {"batch_id": batch_id, "filename": file.filename, "rows_parsed": len(rows), "inserted": inserted}

# ─── Scout ────────────────────────────────────────────────

@app.post("/api/scout/run")
async def run_scouts(limit: int = 20, tier: Optional[str] = None, rescout: bool = False, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    """Scout leads for buying signals.
    - Default: scouts unscouted leads (buying_signals IS NULL)
    - tier=warm: scouts all warm-tier leads
    - rescout=true: re-scouts even if they already have signals
    """
    async with httpx.AsyncClient(timeout=30) as client:
        run_resp = await client.post(f"{SUPABASE_URL}/rest/v1/scout_runs", headers=HEADERS_SB,
                                     json={"run_type": f"manual_tier={tier or 'unscouted'}", "status": "running"})
        run_resp.raise_for_status()
        run_id = run_resp.json()[0]["id"]

        # Build query based on filters
        url = f"{SUPABASE_URL}/rest/v1/leads?limit={limit}&order=rank_score.desc.nullslast&user_id=eq.{user['user_id']}"
        if tier:
            url += f"&qualification_tier=eq.{tier}"
        if not rescout and not tier:
            url += "&buying_signals=is.null"

        leads_resp = await client.get(url, headers=HEADERS_SB)
        leads_resp.raise_for_status()
        leads = leads_resp.json()

    if not leads:
        return {"run_id": run_id, "scouted": 0, "message": f"No leads to scout (tier={tier}, rescout={rescout})"}

    updated = 0
    errors = []
    for i in range(0, len(leads), 5):
        batch = leads[i:i+5]
        lens = await _get_user_sales_lens(user["user_id"])
        results = await asyncio.gather(*[ScoutAgent.scout_lead(l, sales_lens=lens) for l in batch], return_exceptions=True)
        async with httpx.AsyncClient(timeout=30) as client:
            for lead, result in zip(batch, results):
                if isinstance(result, Exception):
                    errors.append(f"{lead.get('full_name','?')}: {str(result)[:100]}")
                    continue
                now = datetime.now(timezone.utc).isoformat()
                await client.patch(f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}", headers=HEADERS_SB,
                    json={**result, "buying_signals_updated_at": now,
                          "web_signals_updated_at": now, "linkedin_summary_updated_at": now, "updated_at": now})
                updated += 1

    # Auto-rank scouted leads
    ranked = 0
    async with httpx.AsyncClient(timeout=30) as client:
        for lead in leads[:updated]:
            lead_resp = await client.get(
                f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}&select=id,buying_signals,web_signals,linkedin_summary,title,company_industry,company_size,company_description",
                headers=HEADERS_SB)
            if lead_resp.status_code == 200 and lead_resp.json():
                ld = lead_resp.json()[0]
                analysis = CERATABridge.analyze(
                    buying_signals=ld.get("buying_signals",""), web_signals=ld.get("web_signals",""),
                    linkedin_summary=ld.get("linkedin_summary",""), title=ld.get("title",""),
                    company_industry=ld.get("company_industry",""),
                    company_size=ld.get("company_size") or 0,
                    company_description=ld.get("company_description",""))
                now = datetime.now(timezone.utc).isoformat()
                await client.patch(f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}", headers=HEADERS_SB,
                    json={**analysis, "ranked_at": now, "updated_at": now})
                ranked += 1

    # Update run record
    async with httpx.AsyncClient(timeout=30) as client:
        await client.patch(f"{SUPABASE_URL}/rest/v1/scout_runs?id=eq.{run_id}", headers=HEADERS_SB,
            json={"status": "complete", "completed_at": datetime.now(timezone.utc).isoformat(),
                  "leads_scouted": len(leads), "leads_updated": updated,
                  "errors": "; ".join(errors[:10]) if errors else None})

    return {"run_id": run_id, "scouted": len(leads), "updated": updated, "ranked": ranked, "errors": len(errors)}

# ─── Rank ─────────────────────────────────────────────────

@app.post("/api/rank/run")
async def run_ranking(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    icp = await _get_user_icp(user["user_id"])
    async with httpx.AsyncClient(timeout=30) as client:
        run_resp = await client.post(f"{SUPABASE_URL}/rest/v1/rank_runs", headers=HEADERS_SB,
                                     json={"status": "running"})
        run_resp.raise_for_status()
        run_id = run_resp.json()[0]["id"]

        leads_resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/leads?buying_signals=not.is.null&user_id=eq.{user['user_id']}&select=id,buying_signals,web_signals,linkedin_summary,title,company_industry,company_size,company_size_range,company_description,region,company_revenue&limit=1000",
            headers=HEADERS_SB)
        leads_resp.raise_for_status()
        leads = leads_resp.json()

    if not leads:
        return {"run_id": run_id, "ranked": 0}

    tiers = {}
    now = datetime.now(timezone.utc).isoformat()
    async with httpx.AsyncClient(timeout=30) as client:
        for lead in leads:
            a = CERATABridge.analyze(
                buying_signals=lead.get("buying_signals",""), web_signals=lead.get("web_signals",""),
                linkedin_summary=lead.get("linkedin_summary",""), title=lead.get("title",""),
                company_industry=lead.get("company_industry",""),
                company_size=lead.get("company_size") or 0,
                company_description=lead.get("company_description",""),
                icp=icp, lead_data=lead)
            await client.patch(f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}", headers=HEADERS_SB,
                json={**a, "ranked_at": now, "updated_at": now})
            tiers[a["qualification_tier"]] = tiers.get(a["qualification_tier"], 0) + 1

    # Assign positions
    async with httpx.AsyncClient(timeout=30) as client:
        sorted_resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/leads?rank_score=not.is.null&order=rank_score.desc&select=id&limit=1000",
            headers=HEADERS_SB)
        for pos, lead in enumerate(sorted_resp.json(), 1):
            await client.patch(f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}", headers=HEADERS_SB,
                json={"rank_position": pos})

        await client.patch(f"{SUPABASE_URL}/rest/v1/rank_runs?id=eq.{run_id}", headers=HEADERS_SB,
            json={"status": "complete", "completed_at": datetime.now(timezone.utc).isoformat(),
                  "leads_ranked": len(leads), "tier_counts": tiers})

    return {"run_id": run_id, "ranked": len(leads), "tier_counts": tiers}

# ─── Chat (Tool-Using Agent) ───────────────────────────

class ChatRequest(BaseModel):
    message: str
    history: List[Dict[str, str]] = []

CHAT_SYSTEM_BASE = """You are the Rose Glass Sales Intelligence Agent. You have REAL tools to query and modify the leads database.

IMPORTANT: When the user asks you to update a lead, scout a lead, or search for leads, you MUST use the tools provided. 
Do NOT just describe what you would do — actually call the tools.

Dimensions (CERATA framework):
- Ψ (psi_intent): Intent coherence (0-1)
- ρ (rho_authority): Decision authority (0-1)
- q (q_optimized): Bio-optimized urgency (0-1)
- f (f_fit): Ecosystem fit (0-1)
- coherence_score: Overall signal (0-4)
- qualification_tier: hot/warm/cold/disqualified

Be direct. Use dimensional analysis to explain reasoning.

EMAIL DRAFTING RULES — when asked to draft emails:
- Write like a human. No marketing speak. No "I hope this email finds you well."
- Short. Max 120 words. One specific reference to them. No bullet points.
- Never say "streamline", "leverage", "cutting-edge", "game-changer", "transform", or "revolutionize."
- End with a specific low-commitment ask."""


def _build_chat_system(sales_lens: dict = None) -> str:
    """Build the chat system prompt with user's sales lens context."""
    base = CHAT_SYSTEM_BASE
    if not sales_lens or not sales_lens.get("product_name"):
        return base
    
    lens_context = "\n\nUSER'S SALES CONTEXT — frame ALL advice through this lens:"
    lens_context += f"\nProduct: {sales_lens.get('product_name', '')}"
    if sales_lens.get('product_description'):
        lens_context += f"\nWhat it is: {sales_lens['product_description']}"
    if sales_lens.get('value_props'):
        lens_context += f"\nValue props: {', '.join(sales_lens['value_props'])}"
    if sales_lens.get('not_this'):
        lens_context += f"\nCRITICAL — NOT this: {', '.join(sales_lens['not_this'])}"
    if sales_lens.get('industry_terms'):
        lens_context += f"\nUse these terms: {', '.join(sales_lens['industry_terms'])}"
    if sales_lens.get('tone'):
        lens_context += f"\nTone: {sales_lens['tone']}"
    lens_context += "\n\nALWAYS frame buying signals, recommendations, and outreach in terms of the user's product — NOT generic software categories."
    
    return base + lens_context

CHAT_TOOLS = [
    {
        "name": "search_leads",
        "description": "Search leads by name, company, tier, or region. Returns matching leads with all dimensional data.",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Name, company, or keyword to search"},
                "tier": {"type": "string", "description": "Filter by tier: hot, warm, cold, disqualified"},
                "limit": {"type": "integer", "description": "Max results (default 25)"},
            },
        },
    },
    {
        "name": "update_lead",
        "description": "Update a lead's fields. Use this to write notes, change status, record outreach, or save buying signals.",
        "input_schema": {
            "type": "object",
            "properties": {
                "lead_id": {"type": "string", "description": "UUID of the lead to update"},
                "buying_signals": {"type": "string", "description": "Buying signals text to write"},
                "user_notes": {"type": "string", "description": "User notes to add"},
                "user_status": {"type": "string", "description": "Status: new, contacted, qualified, nurture, dead"},
                "user_rating": {"type": "integer", "description": "1-5 star rating"},
                "outreach_status": {"type": "string", "description": "none, emailed, called, meeting_scheduled, proposal_sent"},
                "qualification_tier": {"type": "string", "description": "hot, warm, cold, disqualified — use to reclassify a lead"},
            },
            "required": ["lead_id"],
        },
    },
    {
        "name": "scout_lead",
        "description": "Run a web search scout on a specific lead to find buying signals. Writes results directly to the database.",
        "input_schema": {
            "type": "object",
            "properties": {
                "lead_id": {"type": "string", "description": "UUID of the lead to scout"},
            },
            "required": ["lead_id"],
        },
    },
    {
        "name": "scout_all",
        "description": "Scout ALL unscouted leads for this user. Triggers scouting for all leads that haven't been researched yet. Use when the user says 'scout my leads' or 'scout everything' or 'scout the unscored leads'.",
        "input_schema": {"type": "object", "properties": {"limit": {"type": "integer", "description": "Max leads to scout (default 50)"}}, "required": []},
    },
    {
        "name": "rescout_all",
        "description": "Re-scout ALL leads that have already been scouted, using the user's current Sales Lens. Use when the user says 'rescout my leads', 're-research', or 'scout again through my lens'. This overwrites existing buying signals with fresh research framed through the user's product lens.",
        "input_schema": {"type": "object", "properties": {"limit": {"type": "integer", "description": "Max leads to rescout (default 50)"}}, "required": []},
    },
    {
        "name": "draft_email",
        "description": """Draft a follow-up email for a lead. ALWAYS use this tool when the user asks to draft, write, or compose an email. This tool does NOT draft immediately — it first asks the user essential questions to write from real context, not assumptions. The tool returns questions for the user to answer. Once they answer, call draft_email again with their answers to generate the email.""",
        "input_schema": {
            "type": "object",
            "properties": {
                "lead_name": {"type": "string", "description": "The lead's name"},
                "context": {"type": "string", "description": "Any known context: what happened on the call, what they're interested in, objections raised"},
                "user_answers": {"type": "string", "description": "The user's answers to the intake questions. Leave empty on first call — the tool will return questions. On second call, paste the user's answers here."},
            },
            "required": ["lead_name"],
        },
    },
    {
        "name": "rank_lead",
        "description": "Re-run CERATA dimensional analysis on a lead after new signals are added.",
        "input_schema": {
            "type": "object",
            "properties": {
                "lead_id": {"type": "string", "description": "UUID of the lead to re-rank"},
            },
            "required": ["lead_id"],
        },
    },
]


async def _exec_tool(name: str, inp: Dict, user_id: str = None) -> str:
    """Execute a chat tool and return result as string."""
    try:
        if name == "search_leads":
            query = inp.get("query", "")
            tier = inp.get("tier")
            limit = inp.get("limit", 25)
            url = f"{SUPABASE_URL}/rest/v1/leads?order=rank_score.desc.nullslast&limit={limit}"
            url += "&select=id,full_name,title,company,company_industry,region,locality,rank_score,coherence_score,qualification_tier,psi_intent,rho_authority,q_urgency,q_optimized,f_fit,dimensional_fractures,buying_signals,user_notes,user_status,outreach_status,email,phone_number1,mobile_phone1,linkedin_profile_url,company_domain,company_size_range,company_revenue"
            url += f"&user_id=eq.{user['user_id']}"
            if tier:
                url += f"&qualification_tier=eq.{tier}"
            if query:
                url += f"&or=(full_name.ilike.%25{query}%25,company.ilike.%25{query}%25,region.ilike.%25{query}%25)"
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.get(url, headers=HEADERS_SB)
                resp.raise_for_status()
                leads = resp.json()
            return json.dumps(leads, indent=1)

        elif name == "update_lead":
            lead_id = inp.pop("lead_id")
            data = {k: v for k, v in inp.items() if v is not None}
            now = datetime.now(timezone.utc).isoformat()
            data["updated_at"] = now
            if "buying_signals" in data:
                data["buying_signals_updated_at"] = now
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.patch(
                    f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead_id}&user_id=eq.{user_id}",
                    headers=HEADERS_SB, json=data)
                if resp.status_code >= 400:
                    return f"Error updating lead: {resp.text[:200]}"
                result = resp.json()
            return f"Updated lead {lead_id}: {list(data.keys())}"

        elif name == "scout_lead":
            lead_id = inp["lead_id"]
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.get(
                    f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead_id}",
                    headers=HEADERS_SB)
                leads = resp.json()
                if not leads:
                    return "Lead not found"
                lead = leads[0]
            lens = await _get_user_sales_lens(user_id)
            signals = await ScoutAgent.scout_lead(lead, sales_lens=lens)
            now = datetime.now(timezone.utc).isoformat()
            async with httpx.AsyncClient(timeout=30) as client:
                await client.patch(
                    f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead_id}",
                    headers=HEADERS_SB,
                    json={**signals, "buying_signals_updated_at": now,
                          "web_signals_updated_at": now, "updated_at": now})
            return f"Scouted {lead['full_name']} at {lead['company']}. Signals written to database.\nSignals: {signals['buying_signals'][:500]}"

        elif name == "scout_all":
            limit = inp.get("limit", 50)
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.get(
                    f"{SUPABASE_URL}/rest/v1/leads?buying_signals=is.null&user_id=eq.{user_id}&order=rank_score.desc.nullslast&limit={limit}",
                    headers=HEADERS_SB)
                unscouted = resp.json() if resp.status_code == 200 else []
            if not unscouted:
                return "All leads are already scouted."
            # Launch scouting as background task so chat doesn't block
            asyncio.create_task(_scout_user(user_id, min(limit, len(unscouted))))
            return f"Scouting {len(unscouted)} unscouted leads in the background. They will appear with buying signals and scores within a few minutes. Refresh the Leads tab to see progress."

        elif name == "rescout_all":
            limit = inp.get("limit", 50)
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.get(
                    f"{SUPABASE_URL}/rest/v1/leads?buying_signals=not.is.null&user_id=eq.{user_id}&order=coherence_score.desc&limit={limit}&select=id",
                    headers=HEADERS_SB)
                leads_to_rescout = resp.json() if resp.status_code == 200 else []
            if not leads_to_rescout:
                return "No scouted leads found to re-scout."
            # Clear their buying signals so the background worker picks them up
            async with httpx.AsyncClient(timeout=30) as client:
                for lead in leads_to_rescout:
                    await client.patch(
                        f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}&user_id=eq.{user_id}",
                        headers=HEADERS_SB,
                        json={"buying_signals": None, "rank_score": 99.0, "updated_at": datetime.now(timezone.utc).isoformat()})
            return f"Queued {len(leads_to_rescout)} leads for re-scouting through your Sales Lens. The background worker will process them with your product context. Check the Leads tab in a few minutes to see updated intel."

        elif name == "draft_email":
            lead_name = inp.get("lead_name", "")
            context = inp.get("context", "")
            user_answers = inp.get("user_answers", "")
            
            if not user_answers:
                # First call — return intake questions
                return f"""Before I draft this email, I need a few things from you:

1. **What happened?** — One sentence on what you discussed or what prompted this email.
2. **What did they say they want?** — Their specific interest, concern, or request.
3. **What's the next step?** — What are you proposing? (demo, contract, another call, info send)
4. **Your pricing** — What's the actual cost? (Don't want me making up numbers.)
5. **Anything to avoid?** — Topics, competitors, or framing they wouldn't respond to.

Answer these and I'll draft something that sounds like you wrote it."""
            else:
                # Second call — draft from real answers
                # Get lead data for context
                async with httpx.AsyncClient(timeout=30) as client:
                    resp = await client.get(
                        f"{SUPABASE_URL}/rest/v1/leads?full_name=ilike.%25{lead_name}%25&user_id=eq.{user_id}&limit=1",
                        headers=HEADERS_SB)
                    leads = resp.json() if resp.status_code == 200 else []
                lead = leads[0] if leads else {}
                
                # Get sales lens
                lens = await _get_user_sales_lens(user_id)
                
                prompt = f"""Draft a short follow-up email based on the user's REAL answers below. 

LEAD: {lead.get('full_name', lead_name)} at {lead.get('company', 'their company')}
TITLE: {lead.get('title', '')}

USER'S ANSWERS TO INTAKE QUESTIONS:
{user_answers}

ADDITIONAL CONTEXT: {context}

PRODUCT BEING SOLD: {lens.get('product_name', '')} — {lens.get('product_description', '')}
NEVER POSITION AS: {', '.join(lens.get('not_this', []))}

RULES:
- Under 100 words. No fluff.
- Reference ONE specific thing from the conversation — the thing THEY said, not what you researched.
- Use the pricing THEY gave you. Never invent pricing.
- Sound like a human who just got off the phone, not a marketing team.
- ABSOLUTELY NO markdown formatting. No ** for bold. No bullet points. No dashes as lists. Plain text only.
- No "I hope this finds you well." No "It was great chatting." No "I wanted to follow up." No "Great talking."
- Start with their name and get to the point in the first sentence.
- End with one specific action and a specific time.
- Sign with just the user's first name. No title, no company, no signature block.
- The subject line should be lowercase and casual, like a real person wrote it. No title case.

Output the subject line on the first line (no "Subject:" prefix), then a blank line, then the email body. Nothing else. No labels. No commentary."""

                async with httpx.AsyncClient(timeout=60) as client:
                    resp = await client.post(
                        "https://api.anthropic.com/v1/messages",
                        headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01", "Content-Type": "application/json"},
                        json={"model": "claude-sonnet-4-20250514", "max_tokens": 512, "messages": [{"role": "user", "content": prompt}]})
                    data = resp.json()
                    email_text = "\n".join(b["text"] for b in data.get("content", []) if b.get("type") == "text")
                
                return email_text or "Could not generate email."

        elif name == "rank_lead":
            lead_id = inp["lead_id"]
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.get(
                    f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead_id}&select=id,buying_signals,web_signals,linkedin_summary,title,company_industry,company_size,company_description",
                    headers=HEADERS_SB)
                leads = resp.json()
                if not leads:
                    return "Lead not found"
                lead = leads[0]
            analysis = CERATABridge.analyze(
                buying_signals=lead.get("buying_signals", ""),
                web_signals=lead.get("web_signals", ""),
                linkedin_summary=lead.get("linkedin_summary", ""),
                title=lead.get("title", ""),
                company_industry=lead.get("company_industry", ""),
                company_size=lead.get("company_size") or 0,
                company_description=lead.get("company_description", ""))
            now = datetime.now(timezone.utc).isoformat()
            async with httpx.AsyncClient(timeout=30) as client:
                await client.patch(
                    f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead_id}",
                    headers=HEADERS_SB,
                    json={**analysis, "ranked_at": now, "updated_at": now})
            return f"Re-ranked: {json.dumps(analysis)}"

        return f"Unknown tool: {name}"
    except Exception as e:
        return f"Tool error: {str(e)[:300]}"


@app.post("/api/chat")
async def chat(req: ChatRequest, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    # Build context
    async with httpx.AsyncClient(timeout=30) as client:
        stats_resp = await client.get(f"{SUPABASE_URL}/rest/v1/leads?select=qualification_tier&limit=1000&user_id=eq.{user['user_id']}", headers=HEADERS_SB)
        all_tiers = {}
        for l in stats_resp.json():
            t = l.get("qualification_tier") or "unscored"
            all_tiers[t] = all_tiers.get(t, 0) + 1

        top_resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/leads?order=rank_score.desc.nullslast&limit=25&user_id=eq.{user['user_id']}&select=id,full_name,title,company,region,locality,rank_score,coherence_score,qualification_tier,psi_intent,rho_authority,q_optimized,f_fit,dimensional_fractures,buying_signals,user_notes,user_status,email,phone_number1,mobile_phone1,linkedin_profile_url,company_domain",
            headers=HEADERS_SB)
        top_leads = top_resp.json()

    # Trim top leads context to prevent overflow
    top_leads_slim = [{"full_name": l.get("full_name"), "title": l.get("title"), "company": l.get("company"),
                        "coherence_score": l.get("coherence_score"), "qualification_tier": l.get("qualification_tier"),
                        "email": l.get("email")} for l in top_leads[:15]]
    context = f"DB: {sum(all_tiers.values())} leads, tiers: {json.dumps(all_tiers)}\nTOP 15:\n{json.dumps(top_leads_slim, indent=1)}"
    messages = [*req.history[-6:], {"role": "user", "content": f"{context}\n\nUser: {req.message}"}]

    # Fetch sales lens and build system prompt
    sales_lens = await _get_user_sales_lens(user["user_id"])
    chat_system = _build_chat_system(sales_lens)

    # Agentic loop — keep calling until no more tool_use
    max_iterations = 5
    for _ in range(max_iterations):
        async with httpx.AsyncClient(timeout=120) as client:
            resp = await client.post("https://api.anthropic.com/v1/messages",
                headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01", "Content-Type": "application/json"},
                json={"model": "claude-sonnet-4-20250514", "max_tokens": 4096, "system": chat_system,
                      "tools": CHAT_TOOLS, "messages": messages})
            resp.raise_for_status()
            data = resp.json()

        # Collect response blocks
        assistant_content = data.get("content", [])
        messages.append({"role": "assistant", "content": assistant_content})

        # Check if there are tool calls
        tool_uses = [b for b in assistant_content if b.get("type") == "tool_use"]
        if not tool_uses:
            break  # No tools called, we're done

        # Execute tools and add results
        tool_results = []
        for tu in tool_uses:
            result = await _exec_tool(tu["name"], tu["input"], user_id=user["user_id"])
            tool_results.append({
                "type": "tool_result",
                "tool_use_id": tu["id"],
                "content": result,
            })
        messages.append({"role": "user", "content": tool_results})

    # Extract final text
    reply_parts = []
    for block in assistant_content:
        if block.get("type") == "text":
            reply_parts.append(block["text"])
    reply = "\n".join(reply_parts) or "Action completed."

    # Save to chat history
    async with httpx.AsyncClient(timeout=30) as client:
        await client.post(f"{SUPABASE_URL}/rest/v1/chat_messages", headers=HEADERS_SB,
            json=[{"role": "user", "content": req.message}, {"role": "assistant", "content": reply}])

    return {"reply": reply, "stats": {"total": sum(all_tiers.values()), "tiers": all_tiers}}

# ─── Focused Chat (Call Mode) ──────────────────────────

class FocusChatRequest(BaseModel):
    message: str
    lead_id: str
    history: List[Dict[str, str]] = []

FOCUS_SYSTEM_BASE = """You are the Rose Glass Sales Intelligence Agent in CALL MODE — focused on a single lead.

The user is on a live call or preparing for one. Your job:
1. Before the call: provide intel, suggested openers, talking points
2. During the call: capture what the user types as call notes, update the lead profile in real time
3. After the call: summarize, draft follow-ups, re-rank

You have tools to update this lead's record. ALWAYS use update_lead to write notes, status changes, 
and new intel that the user shares from the call. Every piece of info from the conversation should be 
persisted — don't just acknowledge, WRITE IT.

When the user shares new info (e.g. "they're hiring 3 people", "uses spreadsheets for intake"), 
immediately call update_lead to append it to buying_signals and user_notes.

After significant new intel, call rank_lead to re-compute dimensions.

Be concise — the user is multitasking on a call. Lead with the actionable insight.

EMAIL DRAFTING RULES — when the user asks you to draft an email:
- Write like a human, not an AI. No marketing speak. No "I hope this email finds you well."
- Short paragraphs. One idea per paragraph. Max 120 words total.
- Reference ONE specific thing about them — not a laundry list of research.
- Sound like a person who actually knows the industry, not a template.
- No bullet points in emails. No bold text. No formatting.
- End with a specific, low-commitment ask. Not "let's hop on a call."
- Never say "streamline", "leverage", "cutting-edge", "game-changer", "transform", or "revolutionize."
- Never start with the recipient's company name + flattery."""

def _build_focus_system(sales_lens: dict = None) -> str:
    """Build the focus chat system prompt with user's sales lens."""
    base = FOCUS_SYSTEM_BASE
    if not sales_lens or not sales_lens.get("product_name"):
        return base
    
    lens = "\n\nUSER'S PRODUCT — frame ALL call advice through this lens:"
    lens += f"\nSelling: {sales_lens.get('product_name', '')} — {sales_lens.get('product_description', '')}"
    if sales_lens.get('not_this'):
        lens += f"\nNEVER position as: {', '.join(sales_lens['not_this'])}"
    if sales_lens.get('value_props'):
        lens += f"\nKey value: {', '.join(sales_lens['value_props'])}"
    if sales_lens.get('tone'):
        lens += f"\nTone: {sales_lens['tone']}"
    
    return base + lens


@app.post("/api/chat/focus")
async def focus_chat(req: FocusChatRequest, authorization: str = Header(None)):
    user = await get_current_user(authorization)

    # Get full lead data
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/leads?id=eq.{req.lead_id}&user_id=eq.{user['user_id']}",
            headers=HEADERS_SB)
        leads = resp.json()
    if not leads:
        raise HTTPException(404, "Lead not found")
    lead = leads[0]

    sales_lens = await _get_user_sales_lens(user["user_id"])
    # Trim lead context to essential fields — full buying_signals can be huge
    lead_summary = {
        "full_name": lead.get("full_name"),
        "title": lead.get("title"),
        "company": lead.get("company"),
        "email": lead.get("email"),
        "phone": lead.get("phone_number1") or lead.get("mobile_phone1"),
        "region": lead.get("region"),
        "company_industry": lead.get("company_industry"),
        "coherence_score": lead.get("coherence_score"),
        "qualification_tier": lead.get("qualification_tier"),
        "psi_intent": lead.get("psi_intent"),
        "rho_authority": lead.get("rho_authority"),
        "q_optimized": lead.get("q_optimized"),
        "f_fit": lead.get("f_fit"),
        "dimensional_fractures": lead.get("dimensional_fractures"),
        "buying_signals": (lead.get("buying_signals") or "")[:1500],  # Cap at 1500 chars
        "user_notes": (lead.get("user_notes") or "")[:500],
    }
    context = f"FOCUSED LEAD:\n{json.dumps(lead_summary, indent=1)}"
    # Only keep last 6 history messages to prevent context overflow
    recent_history = req.history[-6:] if req.history else []
    messages = [*recent_history, {"role": "user", "content": f"{context}\n\nUser: {req.message}"}]

    # Agentic loop with tools
    max_iterations = 5
    assistant_content = []
    for _ in range(max_iterations):
        async with httpx.AsyncClient(timeout=120) as client:
            resp = await client.post("https://api.anthropic.com/v1/messages",
                headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01", "Content-Type": "application/json"},
                json={"model": "claude-sonnet-4-20250514", "max_tokens": 4096, "system": _build_focus_system(sales_lens),
                      "tools": CHAT_TOOLS, "messages": messages})
            resp.raise_for_status()
            data = resp.json()

        assistant_content = data.get("content", [])
        messages.append({"role": "assistant", "content": assistant_content})

        tool_uses = [b for b in assistant_content if b.get("type") == "tool_use"]
        if not tool_uses:
            break

        tool_results = []
        for tu in tool_uses:
            result = await _exec_tool(tu["name"], tu["input"], user_id=user["user_id"])
            tool_results.append({"type": "tool_result", "tool_use_id": tu["id"], "content": result})
        messages.append({"role": "user", "content": tool_results})

    reply = "\n".join(b["text"] for b in assistant_content if b.get("type") == "text") or "Done."

    # Check if any tool calls wrote to the database
    wrote_to_db = any(
        b.get("name") in ("update_lead", "scout_lead", "rank_lead", "scout_all")
        for content_list in messages
        if isinstance(content_list, dict) and isinstance(content_list.get("content"), list)
        for b in content_list["content"]
        if isinstance(b, dict) and b.get("type") == "tool_use"
    )

    # Get updated lead after any tool calls
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/leads?id=eq.{req.lead_id}&user_id=eq.{user['user_id']}",
            headers=HEADERS_SB)
        updated_lead = resp.json()[0] if resp.json() else lead

    return {"reply": reply, "lead": updated_lead, "wrote": wrote_to_db}

# ─── ICP Profile ──────────────────────────────────────────

class ICPProfile(BaseModel):
    target_titles: List[str] = []
    target_industries: List[str] = []
    target_company_size_min: Optional[int] = None
    target_company_size_max: Optional[int] = None
    target_regions: List[str] = []
    target_keywords: List[str] = []
    exclude_keywords: List[str] = []
    notes: str = ""

@app.get("/api/icp")
async def get_icp(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/users?id=eq.{user['user_id']}&select=icp_profile",
            headers=HEADERS_SB)
        data = resp.json()
    return data[0].get("icp_profile") or {} if data else {}

@app.post("/api/icp")
async def save_icp(profile: ICPProfile, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.patch(
            f"{SUPABASE_URL}/rest/v1/users?id=eq.{user['user_id']}",
            headers=HEADERS_SB,
            json={"icp_profile": profile.model_dump()})
    return {"saved": True, "profile": profile.model_dump()}


# ─── Sales Lens ───────────────────────────────────────────

class SalesLens(BaseModel):
    product_name: str = ""
    product_description: str = ""
    value_props: List[str] = []
    not_this: List[str] = []
    industry_terms: List[str] = []
    tone: str = ""

@app.get("/api/sales-lens")
async def get_sales_lens(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/users?id=eq.{user['user_id']}&select=sales_lens",
            headers=HEADERS_SB)
        data = resp.json()
    return data[0].get("sales_lens") or {} if data else {}

@app.post("/api/sales-lens")
async def save_sales_lens(lens: SalesLens, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    async with httpx.AsyncClient(timeout=30) as client:
        await client.patch(
            f"{SUPABASE_URL}/rest/v1/users?id=eq.{user['user_id']}",
            headers=HEADERS_SB,
            json={"sales_lens": lens.model_dump()})
    return {"saved": True, "lens": lens.model_dump()}

# ─── Lead CRUD ────────────────────────────────────────────

@app.get("/api/leads")
async def get_leads(tier: Optional[str] = None, limit: int = 500, offset: int = 0, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    url = f"{SUPABASE_URL}/rest/v1/leads?order=rank_score.desc.nullslast&limit={limit}&offset={offset}&user_id=eq.{user['user_id']}"
    if tier: url += f"&qualification_tier=eq.{tier}"
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(url, headers=HEADERS_SB)
        resp.raise_for_status()
        return resp.json()

@app.get("/api/leads/{lead_id}")
async def get_lead(lead_id: str, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead_id}", headers=HEADERS_SB)
        leads = resp.json()
        if not leads: raise HTTPException(404, "Not found")
        return leads[0]

class LeadUpdate(BaseModel):
    user_notes: Optional[str] = None
    user_status: Optional[str] = None
    user_rating: Optional[int] = None
    outreach_status: Optional[str] = None

@app.patch("/api/leads/{lead_id}")
async def update_lead(lead_id: str, update: LeadUpdate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    data = {k: v for k, v in update.dict().items() if v is not None}
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.patch(f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead_id}", headers=HEADERS_SB, json=data)
        return resp.json()

@app.get("/api/stats")
async def get_stats(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.post(
            f"{SUPABASE_URL}/rest/v1/rpc/get_lead_stats",
            headers=HEADERS_SB,
            json={"p_user_id": user["user_id"]})
        data = resp.json() if resp.status_code == 200 else {}
    # Remove zero-count tiers for cleaner display
    tiers = {k: v for k, v in (data.get("tiers") or {}).items() if v > 0}
    return {"total": data.get("total", 0), "scored": data.get("scored", 0), "tiers": tiers, "avg_coherence": 0}

# ─── Background Scout Scheduler ───────────────────────────

SCOUT_INTERVAL = int(os.environ.get("SCOUT_INTERVAL_SECONDS", "60"))  # 5 min default
SCOUT_BATCH = int(os.environ.get("SCOUT_BATCH_SIZE", "20"))
SCOUT_ENABLED = os.environ.get("SCOUT_ENABLED", "true").lower() == "true"

async def _get_user_icp(user_id: str) -> dict:
    """Fetch the user's ICP profile from Supabase."""
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(
                f"{SUPABASE_URL}/rest/v1/users?id=eq.{user_id}&select=icp_profile",
                headers=HEADERS_SB)
            data = resp.json() if resp.status_code == 200 else []
            return data[0].get("icp_profile") or {} if data else {}
    except:
        return {}


async def _get_user_sales_lens(user_id: str) -> dict:
    """Fetch the user's sales lens from Supabase."""
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(
                f"{SUPABASE_URL}/rest/v1/users?id=eq.{user_id}&select=sales_lens",
                headers=HEADERS_SB)
            data = resp.json() if resp.status_code == 200 else []
            return data[0].get("sales_lens") or {} if data else {}
    except:
        return {}


async def _scout_user(user_id: str, batch_size: int = 5):
    """Scout one batch for a single user. Returns count scouted."""
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(
            f"{SUPABASE_URL}/rest/v1/leads?buying_signals=is.null&user_id=eq.{user_id}&order=rank_score.desc.nullslast&limit={batch_size}",
            headers=HEADERS_SB)
        leads = resp.json() if resp.status_code == 200 else []

    if not leads:
        return 0
    
    # Fetch ICP and Sales Lens once per batch
    icp = await _get_user_icp(user_id)
    sales_lens = await _get_user_sales_lens(user_id)

    count = 0
    for lead in leads:
        try:
            signals = await ScoutAgent.scout_lead(lead, sales_lens=sales_lens)
            now = datetime.now(timezone.utc).isoformat()
            async with httpx.AsyncClient(timeout=30) as client:
                await client.patch(
                    f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}",
                    headers=HEADERS_SB,
                    json={**signals, "buying_signals_updated_at": now,
                          "web_signals_updated_at": now, "updated_at": now})

                # Auto-rank
                ld_resp = await client.get(
                    f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}&select=id,buying_signals,web_signals,linkedin_summary,title,company_industry,company_size,company_description",
                    headers=HEADERS_SB)
                if ld_resp.json():
                    ld = ld_resp.json()[0]
                    analysis = CERATABridge.analyze(
                        buying_signals=ld.get("buying_signals",""),
                        web_signals=ld.get("web_signals",""),
                        linkedin_summary=ld.get("linkedin_summary",""),
                        title=ld.get("title",""),
                        company_industry=ld.get("company_industry",""),
                        company_size=ld.get("company_size") or 0,
                        company_description=ld.get("company_description",""),
                        icp=icp, lead_data=ld)
                    await client.patch(
                        f"{SUPABASE_URL}/rest/v1/leads?id=eq.{lead['id']}",
                        headers=HEADERS_SB,
                        json={**analysis, "ranked_at": now, "updated_at": now})

            logger.info(f"Scout: {lead.get('full_name','')} at {lead.get('company','')} [user={user_id[:8]}]")
            count += 1
        except Exception as e:
            logger.error(f"Scout failed: {lead.get('full_name','')}: {e}")
    return count


async def _background_scout_loop():
    """Per-user concurrent scouting. Each user gets their own scout lane."""
    await asyncio.sleep(30)
    logger.info(f"Background scout started: batch={SCOUT_BATCH}, interval={SCOUT_INTERVAL}s, model={SCOUT_MODEL}")

    while True:
        try:
            # Get all users with unscouted leads
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.get(
                    f"{SUPABASE_URL}/rest/v1/leads?buying_signals=is.null&select=user_id&limit=200",
                    headers=HEADERS_SB)
                rows = resp.json() if resp.status_code == 200 else []

            user_ids = list(set(r["user_id"] for r in rows if r.get("user_id")))

            if not user_ids:
                logger.info("Background scout: all leads scouted. Sleeping 1 hour.")
                await asyncio.sleep(3600)
                continue

            # Scout one batch per user CONCURRENTLY
            logger.info(f"Background scout: {len(user_ids)} users with unscouted leads")
            tasks = [_scout_user(uid, SCOUT_BATCH) for uid in user_ids]
            results = await asyncio.gather(*tasks, return_exceptions=True)

            total = sum(r for r in results if isinstance(r, int))
            logger.info(f"Background scout cycle: {total} leads scouted across {len(user_ids)} users")

        except Exception as e:
            logger.error(f"Background scout error: {e}")

        await asyncio.sleep(SCOUT_INTERVAL)


@app.on_event("startup")
async def startup():
    if SCOUT_ENABLED and ANTHROPIC_API_KEY:
        asyncio.create_task(_background_scout_loop())
        logger.info("Background scout scheduler registered")
    else:
        logger.info("Background scout disabled (set SCOUT_ENABLED=true)")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=PORT)
